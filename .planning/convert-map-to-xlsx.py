#!/usr/bin/env python3
"""
Convert UI-DB-CONNECTION-MAP.md → UI-DB-CONNECTION-MAP.xlsx
One sheet per page/section + Summary sheet.
Run: python3 .planning/convert-map-to-xlsx.py
"""

import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MD_PATH = os.path.join(os.path.dirname(__file__), "UI-DB-CONNECTION-MAP.md")
XLSX_PATH = os.path.join(os.path.dirname(__file__), "UI-DB-CONNECTION-MAP.xlsx")

# Status emoji → text mapping for clean Excel cells
STATUS_MAP = {
    "\u2705": "READY",
    "\u26a0\ufe0f": "PENDING",
    "\u26a0": "PENDING",
    "\U0001f534": "NO DB COLUMN",
    "\U0001f7e1": "NOT IN UI",
    "\U0001f501": "DERIVED",
}

# Colors for status
STATUS_FILLS = {
    "READY": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),        # green
    "PENDING": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),       # amber
    "NO DB COLUMN": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),   # red
    "NOT IN UI": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),      # light yellow
    "DERIVED": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),       # blue
}

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FONT = Font(bold=True, size=11, color="1F4E79")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def clean_status(text: str) -> str:
    """Extract status label from a cell that may contain emoji + text."""
    text = text.strip()
    for emoji, label in STATUS_MAP.items():
        if emoji in text:
            # Return the label plus any extra text after it
            remainder = text.replace(emoji, "").strip()
            if remainder:
                return f"{label}: {remainder}"
            return label
    return text


def parse_md_sections(md_text: str):
    """Parse the markdown into numbered sections with their tables."""
    sections = []
    current_section = None
    current_subsection = None
    current_table = None
    table_headers = None

    lines = md_text.split("\n")
    i = 0
    while i < len(lines):
        line = lines[i]

        # H2 = major section
        m = re.match(r"^## (\d+)\.\s+(.+)", line)
        if m:
            if current_section and current_section["tables"]:
                sections.append(current_section)
            current_section = {
                "number": m.group(1),
                "title": m.group(2).strip().split("(")[0].strip(),  # remove (path)
                "path": "",
                "tables": [],
            }
            # Extract path from backticks
            path_m = re.search(r"`([^`]+)`", line)
            if path_m:
                current_section["path"] = path_m.group(1)
            current_subsection = None
            current_table = None
            i += 1
            continue

        # H2 for Summary/Enum sections (no number)
        m2 = re.match(r"^## (Summary|Enum)", line)
        if m2:
            if current_section and current_section["tables"]:
                sections.append(current_section)
            current_section = {
                "number": "S",
                "title": line.lstrip("#").strip(),
                "path": "",
                "tables": [],
            }
            current_subsection = None
            i += 1
            continue

        # H3 = subsection
        if line.startswith("### "):
            current_subsection = line.lstrip("# ").strip()
            current_table = None
            i += 1
            continue

        # Table row
        if line.startswith("|") and current_section:
            cells = [c.strip() for c in line.split("|")[1:-1]]
            if not cells:
                i += 1
                continue

            # Separator row
            if all(re.match(r"^-+$", c) for c in cells):
                i += 1
                continue

            # Header row (if no current table or new header pattern)
            if current_table is None or (table_headers and len(cells) != len(table_headers)):
                current_table = {
                    "subsection": current_subsection or "",
                    "headers": cells,
                    "rows": [],
                }
                table_headers = cells
                current_section["tables"].append(current_table)
                i += 1
                continue

            # Data row
            current_table["rows"].append(cells)
            i += 1
            continue

        # Non-table line resets table context
        if not line.startswith("|") and line.strip() and not line.startswith(">") and not line.startswith("---") and not line.startswith("*"):
            current_table = None
            table_headers = None

        i += 1

    if current_section and current_section["tables"]:
        sections.append(current_section)

    return sections


def write_sheet(ws, section):
    """Write a section's tables to a worksheet."""
    row = 1

    # Section header
    ws.cell(row=row, column=1, value=f"{section['number']}. {section['title']}")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, color="1F4E79")
    if section["path"]:
        ws.cell(row=row, column=2, value=section["path"])
        ws.cell(row=row, column=2).font = Font(italic=True, color="808080", size=10)
    row += 2

    for table in section["tables"]:
        # Subsection header
        if table["subsection"]:
            ws.cell(row=row, column=1, value=table["subsection"])
            ws.cell(row=row, column=1).font = SECTION_FONT
            row += 1

        # Table headers
        for ci, header in enumerate(table["headers"]):
            cell = ws.cell(row=row, column=ci + 1, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
        row += 1

        # Table rows
        for data_row in table["rows"]:
            for ci, val in enumerate(data_row):
                clean_val = clean_status(val) if ci == len(data_row) - 1 else val
                # Remove markdown bold/backticks
                clean_val = re.sub(r"\*\*(.+?)\*\*", r"\1", clean_val)
                clean_val = re.sub(r"`(.+?)`", r"\1", clean_val)
                cell = ws.cell(row=row, column=ci + 1, value=clean_val)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(wrap_text=True, vertical="top")

                # Color-code status column
                if ci == len(data_row) - 1:
                    for status_key, fill in STATUS_FILLS.items():
                        if clean_val.startswith(status_key):
                            cell.fill = fill
                            break
            row += 1

        row += 1  # gap between tables

    # Auto-width columns
    for col_idx in range(1, ws.max_column + 1):
        max_len = 12
        for r in range(1, ws.max_row + 1):
            val = ws.cell(row=r, column=col_idx).value
            if val:
                max_len = max(max_len, min(len(str(val)), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2


def write_instructions_sheet(wb):
    """Create the first tab with legend, instructions, and reading guide."""
    ws = wb.create_sheet(title="Instructions & Legend", index=0)

    # ── Title ──
    row = 1
    cell = ws.cell(row=row, column=1, value="UI \u2194 Database Connection Map")
    cell.font = Font(bold=True, size=18, color="1F4E79")
    ws.merge_cells("A1:D1")
    row += 1
    cell = ws.cell(row=row, column=1, value="TFS Flight School \u2014 Part 61 Prototype")
    cell.font = Font(italic=True, size=12, color="808080")
    ws.merge_cells("A2:D2")
    row += 2

    # ── Purpose ──
    cell = ws.cell(row=row, column=1, value="What is this document?")
    cell.font = Font(bold=True, size=13, color="1F4E79")
    row += 1
    purpose_lines = [
        "This workbook tracks every connection between the UI prototype and the database schema.",
        "It maps each visible element in the prototype to its corresponding database table and column,",
        "showing what's ready to wire, what's still using mock data, and what gaps exist.",
        "",
        "This is a LIVING DOCUMENT \u2014 it updates automatically when the UI or DB schema changes.",
        "An automated agent scans both codebases and pushes updates via GitHub Actions.",
    ]
    for line in purpose_lines:
        ws.cell(row=row, column=1, value=line).font = Font(size=11)
        row += 1
    row += 1

    # ── How to read ──
    cell = ws.cell(row=row, column=1, value="How to Read Each Tab")
    cell.font = Font(bold=True, size=13, color="1F4E79")
    row += 1
    reading_lines = [
        ("Tab structure:", "Each tab represents one page or major section of the UI prototype."),
        ("Tables within tabs:", "Each table maps a UI subsection (e.g., 'Event Detail Sheet', 'Demographics Tab')."),
        ("Column: UI Element", "The visible field, button, badge, or data point shown in the prototype."),
        ("Column: DB Table.Column", "The database table and column that should provide the data (e.g., users.full_name)."),
        ("Column: Status", "Color-coded connection status \u2014 see Legend below."),
        ("Column: DB Operation", "For action items: INSERT, UPDATE, DELETE, or the query logic needed."),
        ("'DB Columns NOT in UI'", "Tables at the bottom of each tab listing database columns that exist but aren't shown yet."),
    ]
    for label, desc in reading_lines:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=11)
        ws.cell(row=row, column=2, value=desc).font = Font(size=11)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        row += 1
    row += 1

    # ── Status Legend ──
    cell = ws.cell(row=row, column=1, value="Status Legend")
    cell.font = Font(bold=True, size=13, color="1F4E79")
    row += 1

    legend_items = [
        ("READY", STATUS_FILLS["READY"],
         "UI field has a matching DB column. Ready to wire up in code."),
        ("PENDING", STATUS_FILLS["PENDING"],
         "UI field exists but uses hardcoded/mock data. The DB column exists \u2014 connection just needs to be built."),
        ("NO DB COLUMN", STATUS_FILLS["NO DB COLUMN"],
         "UI shows this field but NO database column exists for it yet. Needs schema work or a design decision."),
        ("NOT IN UI", STATUS_FILLS["NOT IN UI"],
         "Database column exists but is NOT shown anywhere in the prototype. May need a UI surface in the future."),
        ("DERIVED", STATUS_FILLS["DERIVED"],
         "Computed or derived value. Requires a query/calculation, not a direct column lookup (e.g., COUNT, SUM, date math)."),
    ]

    # Header
    for ci, h in enumerate(["Status", "Color", "Meaning"]):
        cell = ws.cell(row=row, column=ci + 1, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
    row += 1

    for label, fill, meaning in legend_items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=11)
        ws.cell(row=row, column=1).border = THIN_BORDER
        color_cell = ws.cell(row=row, column=2, value="")
        color_cell.fill = fill
        color_cell.border = THIN_BORDER
        meaning_cell = ws.cell(row=row, column=3, value=meaning)
        meaning_cell.font = Font(size=11)
        meaning_cell.alignment = Alignment(wrap_text=True)
        meaning_cell.border = THIN_BORDER
        row += 1
    row += 1

    # ── Key Concepts ──
    cell = ws.cell(row=row, column=1, value="Key Concepts")
    cell.font = Font(bold=True, size=13, color="1F4E79")
    row += 1
    concepts = [
        ("Part 61 School", "A flight school operating under FAA Part 61 rules (not Part 141). Training is instructor-customized, not FAA-approved curriculum."),
        ("RLS (Row-Level Security)", "Database enforces that users can only see data belonging to their school. Every query is automatically scoped."),
        ("Soft Delete", "Safety-relevant records are never permanently deleted. They get a deleted_at timestamp instead."),
        ("Enrollment Chain", "Student \u2192 Enrollment \u2192 Course Version \u2192 Course. A student is enrolled in a specific version of a course."),
        ("Grade Sheet", "The record of a completed lesson. 'Sealed' means finalized and locked for audit."),
        ("Reservation", "A scheduled event (flight, ground, sim, etc.) linking student + instructor + resource + time."),
        ("FIF Notice", "Flight Information Facility notice \u2014 school-wide announcements that require instructor acknowledgement."),
    ]
    for label, desc in concepts:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=11)
        ws.cell(row=row, column=2, value=desc).font = Font(size=11)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
        row += 1
    row += 1

    # ── Summary Stats ──
    cell = ws.cell(row=row, column=1, value="Quick Stats")
    cell.font = Font(bold=True, size=13, color="1F4E79")
    row += 1
    stats = [
        ("DB tables with UI representation", "~38"),
        ("DB tables with NO UI yet", "9"),
        ("UI elements with NO DB column", "11"),
        ("Enum types fully aligned", "17 of 19"),
        ("Last automated scan", "See individual tab headers"),
    ]
    for label, val in stats:
        ws.cell(row=row, column=1, value=label).font = Font(size=11)
        ws.cell(row=row, column=2, value=val).font = Font(bold=True, size=11)
        row += 1
    row += 1

    # ── Tab Directory ──
    cell = ws.cell(row=row, column=1, value="Tab Directory")
    cell.font = Font(bold=True, size=13, color="1F4E79")
    row += 1

    tabs = [
        ("1. Landing Page", "Public-facing landing page: login, school branding, prospective students"),
        ("2. Instructor Dashboard Shell", "Top bar, profile dropdown, navigation \u2014 shared across all instructor pages"),
        ("3. Schedule Tab", "Instructor's personal schedule: calendar events, add training dialog, blockout times, availability"),
        ("4. Students Tab", "Student roster list with filters, status badges, progress indicators"),
        ("5. Student Detail", "Individual student: demographics, progress, endorsements, tests, stage checks, overrides, sessions"),
        ("6. School Schedule", "School-wide resource grid: aircraft + room time blocks with event interactions"),
        ("7. Live Map", "Real-time aircraft positions, base marker, geofence, weather data"),
        ("8. Maintenance", "Fleet status, due items, squawks, work orders"),
        ("9. Support", "Support ticket form and contact information"),
        ("10. FIF Notice Banner", "Flight Information Facility notices shown across all instructor pages"),
        ("Summary \u2014 Gap Analysis", "Consolidated view: all red gaps, yellow gaps, coverage metrics"),
        ("Enum Alignment Check", "Verifies UI enum values match database enum definitions exactly"),
    ]

    for ci, h in enumerate(["Tab", "Description"]):
        cell = ws.cell(row=row, column=ci + 1, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
    row += 1

    for tab_name, desc in tabs:
        ws.cell(row=row, column=1, value=tab_name).font = Font(bold=True, size=11)
        ws.cell(row=row, column=1).border = THIN_BORDER
        desc_cell = ws.cell(row=row, column=2, value=desc)
        desc_cell.font = Font(size=11)
        desc_cell.alignment = Alignment(wrap_text=True)
        desc_cell.border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        row += 1
    row += 1

    # ── Footer ──
    ws.cell(row=row, column=1, value="This document is maintained by an automated agent. Do not edit directly \u2014 changes will be overwritten on next sync.").font = Font(italic=True, size=10, color="808080")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 20


def main():
    with open(MD_PATH, "r") as f:
        md_text = f.read()

    sections = parse_md_sections(md_text)
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Instructions & Legend as first tab
    write_instructions_sheet(wb)

    for section in sections:
        # Sheet name (max 31 chars, no invalid chars)
        name = f"{section['number']}. {section['title']}"
        # Remove chars invalid in Excel sheet names: : \ / ? * [ ]
        name = re.sub(r'[:\\/\?\*\[\]]', '-', name)
        if len(name) > 31:
            name = name[:28] + "..."
        ws = wb.create_sheet(title=name)
        write_sheet(ws, section)

    # Save
    wb.save(XLSX_PATH)
    print(f"Saved: {XLSX_PATH}")
    print(f"Sheets: {len(wb.sheetnames)}")
    for name in wb.sheetnames:
        print(f"  - {name}")


if __name__ == "__main__":
    main()
